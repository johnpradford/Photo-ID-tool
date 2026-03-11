#!/usr/bin/env python3
"""
Seed the Supabase 'species' table from WAM_species_list_2024.xlsx.

Usage (from repo root):
  pip install openpyxl supabase
  python web/supabase/seed/seed_species.py

Environment variables (or hardcode for a one-time run):
  SUPABASE_URL      - https://llbbexjthcbmgkwkekrc.supabase.co
  SUPABASE_SERVICE_KEY - service_role key (NOT the anon key)
                        Found in: Dashboard → Settings → API → service_role
"""

import os
import sys
import json
import time
from pathlib import Path
import openpyxl

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
SUPABASE_URL = os.getenv("SUPABASE_URL", "https://llbbexjthcbmgkwkekrc.supabase.co")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY", "")  # Must set this

REPO_ROOT = Path(__file__).resolve().parents[3]
WAM_PATH = REPO_ROOT / "WAM_species_list_2024.xlsx"

BATCH_SIZE = 200  # rows per upsert request

# ---------------------------------------------------------------------------
# WAM column → DB field mapping (matches species_db.py COLUMN_MAP)
# ---------------------------------------------------------------------------
COLUMN_MAP = {
    "wam names": "taxon_name",
    "wamnames": "taxon_name",
    "biologic names": "biologic_name",
    "biologicnames": "biologic_name",
    "vernacular": "common_name",
    "class": "class_name",
    "order": "order_name",
    "family": "family",
    "genus": "genus",
    "naturalised": "introduced",
    "epbcconstat": "epbc_con_stat",
    "bcconstat": "bc_con_stat",
    "wapriority": "wa_priority",
    "waconstat": "wa_con_stat",
    # IBSA hidden-sheet aliases
    "taxonname": "taxon_name",
    "commonname": "common_name",
    "familyname": "family",
}


def load_wam_rows(xlsx_path: Path) -> list[dict]:
    """Parse WAM Excel and return list of dicts ready for Supabase upsert."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # Find the right sheet: prefer Sheet1, or the WAM+AFD hidden sheet
    sheet = None
    for name in wb.sheetnames:
        n = name.lower()
        if name == "Sheet1" or ("wam" in n and "afd" in n):
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("No rows found in workbook")

    # Build header → field mapping
    header_row = [str(h).strip().lower().replace(" ", "") if h else "" for h in rows[0]]
    field_map: dict[int, str] = {}
    for idx, header in enumerate(header_row):
        # try exact match, then stripped
        original = str(rows[0][idx]).strip().lower() if rows[0][idx] else ""
        for key, field in COLUMN_MAP.items():
            if original == key or header == key.replace(" ", ""):
                field_map[idx] = field
                break

    print(f"  Detected columns: {field_map}")

    records = []
    for row in rows[1:]:
        rec: dict = {}
        for idx, field in field_map.items():
            val = row[idx]
            if val is not None:
                rec[field] = str(val).strip()
        if not rec.get("taxon_name"):
            continue
        # Ensure required field exists
        records.append(rec)

    wb.close()
    return records


def upsert_batch(url: str, key: str, rows: list[dict]) -> None:
    """POST a batch to Supabase REST API."""
    import urllib.request

    endpoint = f"{url}/rest/v1/species"
    payload = json.dumps(rows).encode()
    req = urllib.request.Request(
        endpoint,
        data=payload,
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates",
        },
        method="POST",
    )
    with urllib.request.urlopen(req) as resp:
        if resp.status not in (200, 201):
            raise RuntimeError(f"HTTP {resp.status}: {resp.read()}")


def main():
    if not SUPABASE_SERVICE_KEY:
        print(
            "ERROR: Set SUPABASE_SERVICE_KEY environment variable to the service_role key.\n"
            "  Dashboard → Settings → API → service_role secret\n"
            "  export SUPABASE_SERVICE_KEY=eyJ..."
        )
        sys.exit(1)

    if not WAM_PATH.exists():
        print(f"ERROR: WAM file not found at {WAM_PATH}")
        sys.exit(1)

    print(f"Loading {WAM_PATH.name} ...")
    rows = load_wam_rows(WAM_PATH)
    print(f"  {len(rows)} species records parsed")

    total = len(rows)
    uploaded = 0
    for i in range(0, total, BATCH_SIZE):
        batch = rows[i : i + BATCH_SIZE]
        print(f"  Upserting rows {i+1}–{min(i+BATCH_SIZE, total)} ...", end=" ", flush=True)
        upsert_batch(SUPABASE_URL, SUPABASE_SERVICE_KEY, batch)
        uploaded += len(batch)
        print("OK")
        time.sleep(0.1)  # be gentle on the API

    print(f"\nDone. {uploaded} species upserted into Supabase.")


if __name__ == "__main__":
    main()
