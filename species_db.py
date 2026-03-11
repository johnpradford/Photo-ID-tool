"""Species database: read WAM 2024 species list, provide fast search index.

Supports two input formats:
1. Standalone WAM species xlsx (Sheet1 with headers like BIOLOGIC NAMES, WAM NAMES, etc.)
2. Hidden sheet inside an IBSA workbook (sheet name contains "WAM" and "AFD")

Actual WAM 2024 columns:
  BIOLOGIC NAMES | WAM NAMES | CLASS | ORDER | FAMILY | GENUS | SPECIES |
  SUBSPECIES | VERNACULAR | NATURALISED | EPBCConStat | BCConStat | WAPriority | WAConStat
"""

import json
import os
from collections import Counter
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import openpyxl


@dataclass
class SpeciesRecord:
    """A species record from the WAM sheet."""
    taxon_name: str = ""        # WAM NAMES (preferred) or BIOLOGIC NAMES
    biologic_name: str = ""     # BIOLOGIC NAMES (Biologic's internal name)
    class_name: str = ""        # CLASS
    order: str = ""             # ORDER
    family_name: str = ""       # FAMILY
    genus: str = ""             # GENUS
    species_epithet: str = ""   # SPECIES
    subspecies: str = ""        # SUBSPECIES
    common_name: str = ""       # VERNACULAR
    introduced: str = ""        # NATURALISED
    epbc_con_stat: str = ""     # EPBCConStat
    bc_con_stat: str = ""       # BCConStat
    wa_priority: str = ""       # WAPriority
    wa_con_stat: str = ""       # WAConStat
    sre_sts: str = ""           # SRE_Sts (if present)
    _search_text: str = ""      # Internal: for fast search

    def to_output_fields(self) -> dict:
        """Return dict mapping to IBSA output column names."""
        return {
            "TaxonName": self.taxon_name,
            "Class": self.class_name,
            "Order": self.order,
            "FamilyName": self.family_name,
            "CommonName": self.common_name,
            "Introduced": self.introduced,
            "EPBCConStat": self.epbc_con_stat,
            "BCConStat": self.bc_con_stat,
            "WAConStat": self.wa_con_stat,
            "SRE_Sts": self.sre_sts,
        }

    def display_text(self) -> str:
        if self.common_name:
            return f"{self.taxon_name} ({self.common_name})"
        return self.taxon_name

    def short_label(self) -> str:
        """Short label for buttons: common name first if available."""
        if self.common_name:
            return f"{self.common_name}\n{self.taxon_name}"
        return self.taxon_name


class SpeciesDB:
    """Searchable species database loaded from WAM Excel sheet."""

    # Mapping from WAM column headers (normalised lowercase) to record fields.
    # Handles both the actual WAM 2024 format and the IBSA hidden-sheet format.
    COLUMN_MAP = {
        # --- Actual WAM 2024 headers ---
        "wam names": "taxon_name",
        "wamnames": "taxon_name",
        "biologic names": "biologic_name",
        "biologicnames": "biologic_name",
        "vernacular": "common_name",
        "naturalised": "introduced",
        "wapriority": "wa_priority",
        # --- Standard/generic headers ---
        "taxonname": "taxon_name",
        "taxon_name": "taxon_name",
        "taxon name": "taxon_name",
        "scientificname": "taxon_name",
        "scientific name": "taxon_name",
        "species name": "taxon_name",
        "class": "class_name",
        "order": "order",
        "family": "family_name",
        "familyname": "family_name",
        "family name": "family_name",
        "genus": "genus",
        "species": "species_epithet",
        "subspecies": "subspecies",
        "commonname": "common_name",
        "common_name": "common_name",
        "common name": "common_name",
        "introduced": "introduced",
        "epbcconstat": "epbc_con_stat",
        "bcconstat": "bc_con_stat",
        "waconstat": "wa_con_stat",
        "sre_sts": "sre_sts",
        "srests": "sre_sts",
    }

    def __init__(self):
        self.species: List[SpeciesRecord] = []
        self._by_taxon: Dict[str, SpeciesRecord] = {}
        self._by_biologic: Dict[str, SpeciesRecord] = {}
        self._by_common: Dict[str, SpeciesRecord] = {}  # lowercase key
        self._loaded = False
        self._source_file = ""

    @property
    def loaded(self) -> bool:
        return self._loaded

    @property
    def count(self) -> int:
        return len(self.species)

    def _build_indices(self):
        """Build lookup indices after loading."""
        self._by_taxon.clear()
        self._by_biologic.clear()
        self._by_common.clear()
        for sp in self.species:
            if sp.taxon_name:
                self._by_taxon[sp.taxon_name] = sp
            if sp.biologic_name and sp.biologic_name != sp.taxon_name:
                self._by_biologic[sp.biologic_name] = sp
            if sp.common_name:
                self._by_common[sp.common_name.lower()] = sp
            # Build search text including all name variants
            sp._search_text = " ".join(filter(None, [
                sp.taxon_name, sp.biologic_name, sp.common_name,
                sp.family_name, sp.order, sp.genus, sp.class_name,
            ])).lower()

    def load_from_workbook(self, workbook_path: str) -> Tuple[bool, str]:
        """Load species from a WAM xlsx file or from a hidden WAM sheet in an IBSA workbook.

        Supports:
        - Standalone WAM species list (Sheet1 with WAM headers)
        - IBSA workbook with hidden "WAM_AFD Names (Fauna - 2024)" sheet
        - Any xlsx with recognisable species columns

        Returns (success: bool, message: str).
        """
        try:
            wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

            # Strategy: try WAM/AFD sheet first, then Sheet1, then first sheet
            sheet = None
            sheet_name = ""

            # 1. Look for a WAM_AFD sheet (IBSA workbook)
            for name in wb.sheetnames:
                if "WAM" in name.upper() and ("AFD" in name.upper() or "FAUNA" in name.upper()):
                    sheet = wb[name]
                    sheet_name = name
                    break

            # 2. Fall back to first sheet
            if sheet is None:
                sheet = wb.active
                sheet_name = wb.sheetnames[0] if wb.sheetnames else "?"

            # Read header row
            rows_iter = sheet.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if not header:
                wb.close()
                return False, "Empty sheet (no header row)"

            # Map column indices to fields
            col_map = {}
            for i, h in enumerate(header):
                if h is None:
                    continue
                h_norm = str(h).strip().lower()
                # Direct match
                if h_norm in self.COLUMN_MAP:
                    col_map[i] = self.COLUMN_MAP[h_norm]
                    continue
                # Stripped match (remove spaces and underscores)
                h_stripped = h_norm.replace(" ", "").replace("_", "")
                for pattern, field in self.COLUMN_MAP.items():
                    if pattern.replace(" ", "").replace("_", "") == h_stripped:
                        col_map[i] = field
                        break

            # Verify we have at least a taxon name column
            mapped_fields = set(col_map.values())
            if "taxon_name" not in mapped_fields and "biologic_name" not in mapped_fields:
                wb.close()
                return False, (
                    f"Cannot find species name column in sheet '{sheet_name}'.\n"
                    f"Headers found: {[h for h in header if h]}\n"
                    f"Expected one of: WAM NAMES, BIOLOGIC NAMES, TaxonName, etc."
                )

            # Read data rows
            self.species = []
            for row in rows_iter:
                rec = SpeciesRecord()
                has_data = False
                for i, field in col_map.items():
                    if i < len(row) and row[i] is not None:
                        val = str(row[i]).strip()
                        if val:
                            setattr(rec, field, val)
                            has_data = True

                if not has_data:
                    continue

                # If we only have biologic_name but no taxon_name, use biologic as taxon
                if not rec.taxon_name and rec.biologic_name:
                    rec.taxon_name = rec.biologic_name

                if rec.taxon_name:
                    self.species.append(rec)

            wb.close()
            self._build_indices()
            self._loaded = True
            self._source_file = workbook_path

            return True, (
                f"Loaded {len(self.species)} species from '{sheet_name}' "
                f"({len(col_map)} columns mapped)"
            )

        except Exception as e:
            return False, f"Error reading workbook: {e}"

    def search(self, query: str, max_results: int = 30) -> List[SpeciesRecord]:
        """Search species by partial name match (scientific, common, family, etc.)."""
        if not query or not query.strip():
            return []

        terms = query.lower().strip().split()

        # Score results: exact prefix matches rank higher
        scored = []
        for sp in self.species:
            if all(t in sp._search_text for t in terms):
                # Boost: taxon name or common name starts with query
                q_lower = query.lower().strip()
                score = 0
                if sp.taxon_name.lower().startswith(q_lower):
                    score = 3
                elif sp.common_name and sp.common_name.lower().startswith(q_lower):
                    score = 2
                elif sp.taxon_name.lower().startswith(terms[0]):
                    score = 1
                scored.append((score, sp))

        scored.sort(key=lambda x: -x[0])
        return [sp for _, sp in scored[:max_results]]

    def get_by_taxon(self, taxon_name: str) -> Optional[SpeciesRecord]:
        """Get exact match by TaxonName (WAM NAMES)."""
        return self._by_taxon.get(taxon_name)

    def get_by_biologic_name(self, name: str) -> Optional[SpeciesRecord]:
        """Get by BIOLOGIC NAMES."""
        return self._by_biologic.get(name) or self._by_taxon.get(name)

    def get_by_common_name(self, name: str) -> Optional[SpeciesRecord]:
        """Get by VERNACULAR (common name), case-insensitive."""
        return self._by_common.get(name.lower())

    def resolve_name(self, name: str) -> Optional[SpeciesRecord]:
        """Resolve a name that could be scientific or common.

        Tries: exact taxon → exact biologic → case-insensitive common name.
        """
        return (self.get_by_taxon(name)
                or self.get_by_biologic_name(name)
                or self.get_by_common_name(name))

    def get_top_species(self, taxon_counts: Dict[str, int], n: int = 20) -> List[SpeciesRecord]:
        """Get top N species by frequency from existing assignments."""
        sorted_taxa = sorted(taxon_counts.items(), key=lambda x: -x[1])[:n]
        results = []
        for taxon, _ in sorted_taxa:
            rec = self.resolve_name(taxon)
            if rec:
                results.append(rec)
        return results

    def load_common_species_file(self, filepath: str) -> Tuple[List[SpeciesRecord], str]:
        """Load a commonly-found species list from an xlsx file.

        The file should have a column of species names (scientific or common).
        Returns the top species by frequency, resolved against the loaded WAM DB.

        Returns (species_list, message).
        """
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active

            # Collect all names
            names = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    names.append(str(row[0]).strip())
            wb.close()

            if not names:
                return [], "No species names found in file"

            # Count frequencies
            counts = Counter(names)

            # Resolve each unique name against the DB
            resolved = []
            unresolved = []
            for name, freq in counts.most_common():
                rec = self.resolve_name(name)
                if rec:
                    resolved.append((rec, freq))
                else:
                    unresolved.append(name)

            species_list = [rec for rec, _ in resolved]

            msg = f"Found {len(names)} records, {len(counts)} unique species, {len(resolved)} matched WAM"
            if unresolved:
                msg += f", {len(unresolved)} unresolved: {', '.join(unresolved[:5])}"
            return species_list, msg

        except Exception as e:
            return [], f"Error reading common species file: {e}"

    def save_preset(self, species_list: List[SpeciesRecord], filepath: str):
        """Save a preset list of species to JSON."""
        data = [{"taxon_name": s.taxon_name, "common_name": s.common_name}
                for s in species_list]
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        with open(filepath, "w") as f:
            json.dump(data, f, indent=2)

    def load_preset(self, filepath: str) -> List[SpeciesRecord]:
        """Load preset species list from JSON, matching to loaded DB."""
        if not os.path.exists(filepath):
            return []
        try:
            with open(filepath) as f:
                data = json.load(f)
            results = []
            for item in data:
                rec = self.resolve_name(item.get("taxon_name", ""))
                if rec:
                    results.append(rec)
            return results
        except (json.JSONDecodeError, KeyError):
            return []
